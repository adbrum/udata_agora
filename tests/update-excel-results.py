#!/usr/bin/env python3
"""
Update Excel test report files with Playwright E2E test results.

Reads the Playwright JSON results from test-results/e2e-results.json and updates
both Excel files (frontend public + backoffice) with pass/fail status, date, and
observations.

Usage:
    python3 tests/update-excel-results.py          (run from frontend/ directory)
    python3 tests/update-excel-results.py --dry-run (preview without writing)
"""

import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# --- Configuration ---

SCRIPT_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = SCRIPT_DIR.parent
RESULTS_FILE = FRONTEND_DIR / "test-results" / "e2e-results.json"

EXCEL_FILES = {
    "frontend": FRONTEND_DIR / ".." / "docs" / "caderno_testes_frontend_publico.xlsx",
    "backoffice": FRONTEND_DIR / ".." / "docs" / "caderno_testes_backoffice_admin.xlsx",
}

SHEET_NAME = "Caderno de Testes"

# Test ID prefixes that belong to each Excel file.
# DS-* is ambiguous (Data Stories in frontend, Datasets in backoffice) and appears in both.
FRONTEND_PREFIXES = {
    "HP",
    "PQ",
    "DL",
    "DD",
    "OL",
    "OD",
    "RL",
    "RD",
    "SD",
    "DS",  # Data Stories
    "TM",
    "MC",
    "NT",
    "AU",
    "PF",
    "DI",
    "PI",
    "NV",
    "RA",
}

BACKOFFICE_PREFIXES = {
    "DS",  # Datasets
    "RU",
    "ORG",
    "API",
    "HV",
    "PO",
    "TP",
    "CR",
    "ED",
    "PM",
    "IA",
    "VL",
    "US",
    "UI",
}

# Prefixes that are unique to one file (used to resolve DS-* ambiguity).
FRONTEND_ONLY_PREFIXES = FRONTEND_PREFIXES - BACKOFFICE_PREFIXES
BACKOFFICE_ONLY_PREFIXES = BACKOFFICE_PREFIXES - FRONTEND_PREFIXES
SHARED_PREFIXES = FRONTEND_PREFIXES & BACKOFFICE_PREFIXES  # {"DS"}


# --- Playwright JSON Parsing ---


def load_results(path: Path) -> dict:
    """Load the Playwright JSON report file."""
    if not path.exists():
        print(f"ERROR: Results file not found: {path}")
        print("Run Playwright tests first with: npx playwright test --reporter=json")
        sys.exit(1)

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_test_id(title: str) -> str | None:
    """Extract test ID (e.g., 'HP-01') from a spec title like 'HP-01: Homepage loads'."""
    match = re.match(r"^([A-Z]+-\d+)", title)
    return match.group(1) if match else None


def get_test_prefix(test_id: str) -> str:
    """Get the prefix from a test ID (e.g., 'HP' from 'HP-01')."""
    return test_id.rsplit("-", 1)[0]


def determine_file_target(test_id: str, suite_path: list[str]) -> str:
    """
    Determine which Excel file a test ID belongs to.

    For shared prefixes (DS-*), uses the suite hierarchy to disambiguate:
    - If the suite path contains backoffice-related keywords -> backoffice
    - Otherwise -> check based on context clues
    """
    prefix = get_test_prefix(test_id)

    if prefix in FRONTEND_ONLY_PREFIXES:
        return "frontend"
    if prefix in BACKOFFICE_ONLY_PREFIXES:
        return "backoffice"

    # Shared prefix (DS-*): disambiguate using suite path context.
    suite_text = " ".join(suite_path).lower()
    backoffice_keywords = [
        "backoffice",
        "admin",
        "dataset",
        "criar",
        "editar",
        "eliminar",
        "publicar",
        "arquivar",
        "listar",
    ]
    frontend_keywords = [
        "data stor",
        "frontend",
        "public",
        "portal",
    ]

    backoffice_score = sum(1 for kw in backoffice_keywords if kw in suite_text)
    frontend_score = sum(1 for kw in frontend_keywords if kw in suite_text)

    if backoffice_score > frontend_score:
        return "backoffice"
    if frontend_score > backoffice_score:
        return "frontend"

    # Default: if the suite file path contains "backoffice", it's backoffice.
    return "backoffice" if "backoffice" in suite_text else "frontend"


def parse_results(data: dict) -> dict[str, dict[str, dict]]:
    """
    Parse Playwright JSON results into a mapping of:
    {file_key: {test_id: {status, error_message}}}
    """
    results = {"frontend": {}, "backoffice": {}}

    def walk_suites(suites: list, path: list[str] | None = None):
        if path is None:
            path = []
        for suite in suites:
            current_path = path + [suite.get("title", "")]

            # Process specs in this suite
            for spec in suite.get("specs", []):
                title = spec.get("title", "")
                test_id = extract_test_id(title)
                if not test_id:
                    continue

                # Determine status from tests array
                status = "skipped"
                error_message = None

                for test in spec.get("tests", []):
                    test_status = test.get("status", "")
                    for result in test.get("results", []):
                        result_status = result.get("status", "")
                        if result_status == "passed":
                            status = "passed"
                        elif result_status == "failed":
                            status = "failed"
                            error_obj = result.get("error", {})
                            if isinstance(error_obj, dict):
                                error_message = error_obj.get("message", "Test failed")
                            else:
                                error_message = (
                                    str(error_obj) if error_obj else "Test failed"
                                )
                        elif (
                            result_status == "skipped"
                            and status != "passed"
                            and status != "failed"
                        ):
                            status = "skipped"

                    # Also check top-level test status
                    if test_status == "expected" and status != "failed":
                        status = "passed"
                    elif test_status == "unexpected":
                        status = "failed"
                        if not error_message:
                            error_message = "Test failed (unexpected status)"
                    elif test_status == "skipped" and status not in (
                        "passed",
                        "failed",
                    ):
                        status = "skipped"

                file_target = determine_file_target(test_id, current_path)

                # Clean ANSI escape codes and illegal characters for Excel
                if error_message:
                    error_message = re.sub(r"\x1b\[[0-9;]*m", "", error_message)
                    error_message = re.sub(
                        r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", error_message
                    )
                    if len(error_message) > 500:
                        error_message = error_message[:497] + "..."

                results[file_target][test_id] = {
                    "status": status,
                    "error_message": error_message,
                }

            # Recurse into nested suites
            walk_suites(suite.get("suites", []), current_path)

    walk_suites(data.get("suites", []))
    return results


# --- Excel Update ---


def find_test_rows(ws) -> dict[str, int]:
    """
    Scan the worksheet and return a mapping of test_id -> row_number
    for all rows where column A matches a test ID pattern.
    """
    test_rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if (
            cell.value
            and isinstance(cell.value, str)
            and re.match(r"^[A-Z]+-\d+$", cell.value)
        ):
            test_rows[cell.value] = cell.row
    return test_rows


def find_section_rows(ws) -> list[dict]:
    """
    Find section header rows and summary table rows.
    Returns a list of dicts with section info for mapping to the summary table.
    """
    sections = []
    current_section = None
    current_section_row = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if (
            cell.value
            and isinstance(cell.value, str)
            and re.match(r"^\d+\.", cell.value)
        ):
            if current_section:
                sections.append(
                    {
                        "name": current_section,
                        "header_row": current_section_row,
                    }
                )
            current_section = cell.value
            current_section_row = cell.row

    if current_section:
        sections.append(
            {
                "name": current_section,
                "header_row": current_section_row,
            }
        )

    return sections


def find_summary_table(ws) -> dict[str, int]:
    """
    Find the summary table and return a mapping of section_name -> row_number.
    The summary table starts with 'Resumo por Secção' and has columns:
    A: Secção, B: Nº Testes, C: Nº OK
    """
    summary_rows = {}
    in_summary = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        a_val = row[0].value
        if a_val and isinstance(a_val, str):
            if "Resumo por Secção" in a_val:
                in_summary = True
                continue
            if in_summary and re.match(r"^\d+\.", a_val):
                summary_rows[a_val.strip()] = row[0].row
            if in_summary and a_val == "TOTAL":
                summary_rows["TOTAL"] = row[0].row

    return summary_rows


def get_section_for_test(
    test_id: str, test_rows: dict[str, int], sections: list[dict]
) -> str | None:
    """Determine which section a test belongs to based on its row position."""
    row_num = test_rows.get(test_id)
    if not row_num:
        return None

    matching_section = None
    for section in sections:
        if section["header_row"] < row_num:
            matching_section = section["name"]
        else:
            break

    return matching_section


def copy_with_timestamp(original_path: Path) -> Path:
    """
    Create a copy of the Excel file with date and time in the name.
    E.g., 'caderno_testes_frontend_publico.xlsx'
       -> 'caderno_testes_frontend_publico_20260322_153045.xlsx'
    """
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    stem = original_path.stem
    suffix = original_path.suffix
    copy_name = f"{stem}_{timestamp}{suffix}"
    copy_path = original_path.parent / copy_name
    shutil.copy2(original_path, copy_path)
    return copy_path


def update_excel(
    file_key: str, file_path: Path, test_results: dict[str, dict], dry_run: bool = False
):
    """
    Copy the original Excel file (adding date/time to the name) and
    fill in test results only in the copy, leaving the original untouched.
    """
    resolved_path = file_path.resolve()
    if not resolved_path.exists():
        print(f"  WARNING: Excel file not found: {resolved_path}")
        return

    # Create timestamped copy
    if not dry_run:
        copy_path = copy_with_timestamp(resolved_path)
        print(f"\n  Original: {resolved_path.name}")
        print(f"  Copy:     {copy_path.name}")
    else:
        copy_path = resolved_path  # dry-run reads original without saving
        print(f"\n  [DRY RUN] Would copy: {resolved_path.name}")

    wb = openpyxl.load_workbook(copy_path)

    if SHEET_NAME not in wb.sheetnames:
        print(f"  WARNING: Sheet '{SHEET_NAME}' not found in {copy_path.name}")
        wb.close()
        return

    ws = wb[SHEET_NAME]
    test_rows = find_test_rows(ws)
    sections = find_section_rows(ws)
    summary_rows = find_summary_table(ws)

    today_str = datetime.now().strftime("%d/%m/%Y")
    updated_count = 0
    not_found_ids = []

    # Track OK counts per section for summary update
    section_ok_counts: dict[str, int] = {}

    # First, count existing OK results for sections that won't be updated
    for test_id, row_num in test_rows.items():
        if test_id not in test_results:
            existing_status = ws.cell(row=row_num, column=6).value
            if existing_status and str(existing_status).strip().upper() == "OK":
                section_name = get_section_for_test(test_id, test_rows, sections)
                if section_name:
                    section_ok_counts[section_name] = (
                        section_ok_counts.get(section_name, 0) + 1
                    )

    # Update test result rows
    for test_id, result in test_results.items():
        if test_id not in test_rows:
            not_found_ids.append(test_id)
            continue

        row_num = test_rows[test_id]
        status = result["status"]
        error_message = result.get("error_message")

        # Column F: OK/NOK
        if status == "passed":
            ok_nok = "OK"
        elif status == "failed":
            ok_nok = "NOK"
        else:
            ok_nok = ""  # skipped: leave empty

        # Column H: Observations
        if status == "passed":
            observation = "Teste automatizado - Playwright"
        elif status == "failed":
            observation = error_message or "Test failed"
        elif status == "skipped":
            observation = "Teste ignorado (skip)"
        else:
            observation = ""

        ws.cell(row=row_num, column=6, value=ok_nok)  # Column F
        ws.cell(row=row_num, column=7, value=today_str)  # Column G
        ws.cell(row=row_num, column=8, value=observation)  # Column H

        # Track OK counts for summary
        if ok_nok == "OK":
            section_name = get_section_for_test(test_id, test_rows, sections)
            if section_name:
                section_ok_counts[section_name] = (
                    section_ok_counts.get(section_name, 0) + 1
                )

        updated_count += 1

        status_symbol = {"passed": "OK", "failed": "NOK", "skipped": "SKIP"}.get(
            status, "?"
        )
        print(f"    {test_id}: {status_symbol}")

    # Update summary table
    summary_updated = 0
    for section_name, ok_count in section_ok_counts.items():
        if section_name in summary_rows:
            summary_row = summary_rows[section_name]
            ws.cell(row=summary_row, column=3, value=ok_count)  # Column C: Nº OK
            summary_updated += 1

    # Update TOTAL row in summary
    if "TOTAL" in summary_rows:
        total_ok = sum(section_ok_counts.values())
        ws.cell(row=summary_rows["TOTAL"], column=3, value=total_ok)

    if not_found_ids:
        print(f"    WARNING: Test IDs not found in Excel: {', '.join(not_found_ids)}")

    print(f"    Updated {updated_count} test(s), {summary_updated} summary section(s)")

    if not dry_run:
        wb.save(copy_path)
        print(f"    Saved: {copy_path.name}")
    else:
        print("    DRY RUN: changes not saved")

    wb.close()


# --- Main ---


def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print("Playwright E2E Results -> Excel Reporter")
    print("=" * 60)

    if dry_run:
        print("\n  [DRY RUN MODE - no files will be modified]\n")

    # Load Playwright results
    print(f"\n  Reading results from: {RESULTS_FILE}")
    data = load_results(RESULTS_FILE)

    # Parse results
    parsed = parse_results(data)

    frontend_count = len(parsed["frontend"])
    backoffice_count = len(parsed["backoffice"])
    total = frontend_count + backoffice_count

    print(
        f"  Found {total} test result(s): {frontend_count} frontend, {backoffice_count} backoffice"
    )

    if total == 0:
        print("\n  No test results with recognized IDs found. Nothing to update.")
        print("  Make sure test titles start with an ID like 'HP-01: ...'")
        sys.exit(0)

    # Summary of statuses
    for key in ("frontend", "backoffice"):
        results = parsed[key]
        if results:
            passed = sum(1 for r in results.values() if r["status"] == "passed")
            failed = sum(1 for r in results.values() if r["status"] == "failed")
            skipped = sum(1 for r in results.values() if r["status"] == "skipped")
            print(f"    {key}: {passed} passed, {failed} failed, {skipped} skipped")

    # Update Excel files
    for file_key, file_path in EXCEL_FILES.items():
        test_results = parsed[file_key]
        if test_results:
            update_excel(file_key, file_path, test_results, dry_run)
        else:
            print(f"\n  {file_key}: no matching results, skipping")

    print("\n" + "=" * 60)
    print("Done.")
    print("=" * 60)


if __name__ == "__main__":
    main()
